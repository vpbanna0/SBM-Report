"""
build-block-pdf.py
Microsoft Graph API se Excel ko PDF convert karta hai.
Excel ki EXACT rendering - bilkul VBA jaisa output.
"""
import os, sys, re, base64, time, shutil, tempfile, subprocess, uuid
from pathlib import Path
from urllib.parse import urlparse, parse_qs, urlencode, quote
import requests

ONEDRIVE_URL = os.environ.get('ONEDRIVE_URL', '')
PDF_ENGINE   = os.environ.get('PDF_ENGINE', 'auto').strip().lower()
PUBLIC_DIR   = Path(__file__).parent.parent / 'public'
DATA_DIR     = PUBLIC_DIR / 'data'
DATA_DIR.mkdir(parents=True, exist_ok=True)

GRAPH_BASE = 'https://graph.microsoft.com/v1.0'
GRAPH_ACCESS_TOKEN   = os.environ.get('GRAPH_ACCESS_TOKEN', '').strip()
MS_GRAPH_TENANT      = os.environ.get('MS_GRAPH_TENANT', 'common').strip() or 'common'
MS_GRAPH_CLIENT_ID   = os.environ.get('MS_GRAPH_CLIENT_ID', '').strip()
MS_GRAPH_CLIENT_SECRET = os.environ.get('MS_GRAPH_CLIENT_SECRET', '').strip()
MS_GRAPH_REFRESH_TOKEN = os.environ.get('MS_GRAPH_REFRESH_TOKEN', '').strip()
DEFAULT_GRAPH_SCOPES = 'https://graph.microsoft.com/Files.ReadWrite offline_access'
MS_GRAPH_SCOPES = os.environ.get('MS_GRAPH_SCOPES', DEFAULT_GRAPH_SCOPES).strip() or DEFAULT_GRAPH_SCOPES

TARGET_SHEETS = [
    'Summary', 'ODF Plus', 'CSC 23-24', 'CSC 24-25', 'CSC 25-26',
    'RRC Updated (4)', 'All IHHL Data Combine', 'Tender Report 25-26',
    'Target AIP 26-27', 'AIP 26-27 Financial', 'Soak Pit 26-27',
    'Compost Pit 26-27', 'Individual assest 26-27'
]

HEADERS = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'accept': 'text/html,application/xhtml+xml,*/*;q=0.8'
}

def norm_sheet(name):
    return re.sub(r'\s+', ' ', str(name or '').strip()).casefold()

TARGET_LOOKUP = {norm_sheet(name): name for name in TARGET_SHEETS}

def safe_file_stem(name):
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', '_', str(name or '').strip())
    cleaned = cleaned.strip(' .')
    return cleaned or 'Block_SBM_Report'

def is_valid_pdf(data):
    return len(data) > 5 and data[:5] == b'%PDF-'

# ═══════════════════════════════════════════════════════
#  STEP 1: Share token se OneDrive item metadata nikalo
# ═══════════════════════════════════════════════════════

def resolve_share_url(url):
    qs = parse_qs(urlparse(url).query)
    redeem = qs.get('redeem', [None])[0]
    if redeem:
        try:
            padded = redeem + '=' * (-len(redeem) % 4)
            return base64.b64decode(padded).decode('utf-8')
        except Exception:
            pass
    return url

def is_valid_zip(data):
    return len(data) > 4 and data[:2] == b'PK'

def get_file_metadata_and_download(url):
    """
    OneDrive share link se:
    1. File metadata (item ID, drive ID) nikalo — Graph API ke liye
    2. Actual xlsx file download karo — PDF naam ke liye
    """
    sharing_url = resolve_share_url(url)
    b64 = base64.b64encode(sharing_url.encode()).decode().rstrip('=').replace('+','-').replace('/','_')
    api_base = f"https://api.onedrive.com/v1.0/shares/u!{b64}"

    session = requests.Session()

    # Metadata fetch
    print("📋 File metadata fetch ho rahi hai...")
    meta = None
    try:
        r = session.get(f"{api_base}/root", timeout=30, headers=HEADERS)
        if r.ok:
            meta = r.json()
            print(f"✅ Metadata mila: {meta.get('name','?')}")
    except Exception as e:
        print(f"⚠️ Metadata fetch failed: {e}")

    # Excel file download (xlsx)
    print("📥 Excel file download ho rahi hai...")
    xlsx_data = None
    try:
        r = session.get(f"{api_base}/root/content", timeout=120, headers=HEADERS, allow_redirects=True)
        if r.ok and is_valid_zip(r.content):
            xlsx_data = r.content
            print(f"✅ Excel downloaded ({len(xlsx_data)//1024} KB)")
        else:
            print(f"⚠️ Direct download failed ({r.status_code}), trying preview scraping...")
            raise Exception("try preview")
    except Exception:
        # Preview scraping fallback
        try:
            from urllib.request import urlopen
            preview_r = session.get(url, headers=HEADERS, allow_redirects=True, timeout=60)
            preview_html = preview_r.text
            match = re.search(
                r'my\.microsoftpersonalcontent\.com[^"\']*download\.aspx\?UniqueId=[^"\']+',
                preview_html, re.IGNORECASE)
            if match:
                signed = match.group(0).replace('\\u0026','&').replace('\\/','/')
                if not signed.startswith('http'):
                    signed = 'https://' + signed
                r2 = session.get(signed, headers=HEADERS, allow_redirects=True, timeout=120)
                if r2.ok and is_valid_zip(r2.content):
                    xlsx_data = r2.content
                    print(f"✅ Excel downloaded via scraping ({len(xlsx_data)//1024} KB)")
        except Exception as e2:
            print(f"⚠️ Scraping failed: {e2}")

    if not xlsx_data:
        raise Exception("Excel file download nahi hui")

    return meta, xlsx_data


# ═══════════════════════════════════════════════════════
#  STEP 2: PDF naam nikalo + sheet names verify karo
# ═══════════════════════════════════════════════════════

def get_pdf_name_and_sheets(xlsx_data):
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True)
    pdf_name = 'Block_SBM_Report'
    sheet_names = wb.sheetnames
    for sname in sheet_names:
        if 'sheet_index' in sname.lower():
            try:
                val = wb[sname]['I2'].value
                if val:
                    pdf_name = safe_file_stem(val)
                    break
            except Exception:
                pass
    wb.close()
    print(f"📄 PDF naam: {pdf_name}")
    print(f"📊 Total sheets in workbook: {len(sheet_names)}")
    # Matched sheets log karo
    matched = [n for n in sheet_names if norm_sheet(n) in TARGET_LOOKUP]
    print(f"✅ Matched target sheets ({len(matched)}): {matched}")
    return pdf_name, sheet_names


# ═══════════════════════════════════════════════════════
#  STEP 3A: Excel Online / Microsoft Graph export
# ═══════════════════════════════════════════════════════

def get_graph_access_token():
    """
    Excel Online rendering ke liye delegated Graph token chahiye.
    GRAPH_ACCESS_TOKEN direct de sakte ho, ya refresh-token flow use kar sakte ho.
    """
    if GRAPH_ACCESS_TOKEN:
        return GRAPH_ACCESS_TOKEN

    if not (MS_GRAPH_CLIENT_ID and MS_GRAPH_REFRESH_TOKEN):
        return ''

    token_url = f"https://login.microsoftonline.com/{MS_GRAPH_TENANT}/oauth2/v2.0/token"
    data = {
        'client_id': MS_GRAPH_CLIENT_ID,
        'grant_type': 'refresh_token',
        'refresh_token': MS_GRAPH_REFRESH_TOKEN,
        'scope': MS_GRAPH_SCOPES,
    }
    if MS_GRAPH_CLIENT_SECRET:
        data['client_secret'] = MS_GRAPH_CLIENT_SECRET

    r = requests.post(token_url, data=data, timeout=45)
    if not r.ok:
        raise RuntimeError(f"Graph token refresh failed: {r.status_code} {r.text[:500]}")
    token = r.json().get('access_token')
    if not token:
        raise RuntimeError("Graph token response me access_token nahi mila")
    return token

def graph_share_id(url):
    sharing_url = resolve_share_url(url)
    encoded = base64.b64encode(sharing_url.encode()).decode()
    encoded = encoded.rstrip('=').replace('/', '_').replace('+', '-')
    return 'u!' + encoded

def graph_request(method, path_or_url, token, expected=(200,), **kwargs):
    url = path_or_url if path_or_url.startswith('http') else GRAPH_BASE + path_or_url
    headers = kwargs.pop('headers', {}) or {}
    headers.setdefault('Authorization', f'Bearer {token}')
    timeout = kwargs.pop('timeout', 120)

    last_response = None
    for attempt in range(5):
        r = requests.request(method, url, headers=headers, timeout=timeout, **kwargs)
        last_response = r
        if r.status_code not in (429, 500, 502, 503, 504):
            break
        delay = int(r.headers.get('Retry-After', '5'))
        print(f"⏳ Graph retry {attempt+1}/5 after {delay}s: {method} {url}")
        time.sleep(delay)

    if last_response.status_code not in expected:
        raise RuntimeError(
            f"Graph {method} failed ({last_response.status_code}): {last_response.text[:700]}"
        )
    return last_response

def resolve_graph_drive_item(token):
    share_id = graph_share_id(ONEDRIVE_URL)
    headers = {'Authorization': f'Bearer {token}', 'Prefer': 'redeemSharingLinkIfNecessary'}
    r = graph_request(
        'GET',
        f"/shares/{share_id}/driveItem?$select=id,name,parentReference",
        token,
        headers=headers,
    )
    item = r.json()
    parent = item.get('parentReference') or {}
    drive_id = parent.get('driveId')
    parent_id = parent.get('id')
    item_id = item.get('id')
    if not (drive_id and parent_id and item_id):
        raise RuntimeError(f"Graph drive item resolve incomplete: {item}")
    print(f"✅ Graph workbook resolve hua: {item.get('name', item_id)}")
    return drive_id, item_id, parent_id

def wait_for_graph_copy(token, monitor_url, drive_id, parent_id, copy_name):
    for attempt in range(90):
        r = graph_request('GET', monitor_url, token, expected=(200, 202), timeout=45)
        try:
            payload = r.json()
        except Exception:
            payload = {}

        status = str(payload.get('status', '')).lower()
        if status in ('completed', 'succeeded') or payload.get('resourceId'):
            resource_id = payload.get('resourceId')
            if not resource_id and payload.get('resourceLocation'):
                match = re.search(r'/items/([^/?]+)', payload['resourceLocation'])
                if match:
                    resource_id = match.group(1)
            if resource_id:
                print("✅ Temp workbook copy ready")
                return resource_id
            break

        if status == 'failed':
            raise RuntimeError(f"Graph copy failed: {payload}")

        if attempt % 10 == 0:
            pct = payload.get('percentageComplete') or payload.get('percentComplete') or '?'
            print(f"⏳ Temp workbook copy wait... {pct}%")
        time.sleep(2)

    # Fallback: monitor response kabhi-kabhi resource id nahi deta.
    r = graph_request(
        'GET',
        f"/drives/{drive_id}/items/{parent_id}/children?$select=id,name&$top=200",
        token,
    )
    for child in r.json().get('value', []):
        if child.get('name') == copy_name:
            print("✅ Temp workbook copy children list se mila")
            return child['id']
    raise RuntimeError("Temp workbook copy complete nahi hui ya locate nahi hui")

def copy_graph_workbook(token, drive_id, item_id, parent_id):
    copy_name = f"__SBM_Block_PDF_{uuid.uuid4().hex}.xlsx"
    body = {
        'parentReference': {'driveId': drive_id, 'id': parent_id},
        'name': copy_name,
    }
    r = graph_request(
        'POST',
        f"/drives/{drive_id}/items/{item_id}/copy",
        token,
        expected=(202,),
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
        json=body,
        timeout=45,
    )
    monitor_url = r.headers.get('Location')
    if not monitor_url:
        raise RuntimeError("Graph copy response me Location header nahi mila")
    return wait_for_graph_copy(token, monitor_url, drive_id, parent_id, copy_name)

def delete_graph_item(token, drive_id, item_id):
    try:
        graph_request('DELETE', f"/drives/{drive_id}/items/{item_id}", token, expected=(204, 404), timeout=45)
        print("🧹 Temp workbook delete ho gaya")
    except Exception as ex:
        print(f"⚠️ Temp workbook delete failed: {ex}")

def list_graph_worksheets(token, drive_id, item_id):
    worksheets = []
    path = f"/drives/{drive_id}/items/{item_id}/workbook/worksheets?$select=id,name,position,visibility"
    while path:
        r = graph_request('GET', path, token)
        payload = r.json()
        worksheets.extend(payload.get('value', []))
        path = payload.get('@odata.nextLink')
    return worksheets

def patch_graph_worksheet(token, drive_id, item_id, worksheet_id, values):
    encoded_id = quote(str(worksheet_id), safe='')
    return graph_request(
        'PATCH',
        f"/drives/{drive_id}/items/{item_id}/workbook/worksheets/{encoded_id}",
        token,
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
        json=values,
    )

def prepare_graph_target_sheets(token, drive_id, item_id):
    worksheets = list_graph_worksheets(token, drive_id, item_id)
    by_norm = {norm_sheet(ws.get('name')): ws for ws in worksheets}

    ordered_targets = []
    missing = []
    for target in TARGET_SHEETS:
        ws = by_norm.get(norm_sheet(target))
        if ws:
            ordered_targets.append(ws)
        else:
            missing.append(target)

    if missing:
        raise RuntimeError(f"Target sheets missing in workbook: {missing}")

    print(f"✅ Graph target sheets ({len(ordered_targets)}): {[ws['name'] for ws in ordered_targets]}")

    # Pehle target sheets visible rakho, phir desired VBA order set karo.
    for ws in ordered_targets:
        patch_graph_worksheet(token, drive_id, item_id, ws['id'], {'visibility': 'Visible'})

    for index, ws in enumerate(ordered_targets):
        patch_graph_worksheet(token, drive_id, item_id, ws['id'], {'position': index})

    target_ids = {ws['id'] for ws in ordered_targets}
    hidden = []
    for ws in worksheets:
        if ws['id'] not in target_ids:
            patch_graph_worksheet(token, drive_id, item_id, ws['id'], {'visibility': 'Hidden'})
            hidden.append(ws.get('name'))

    print(f"🙈 Graph hidden non-target sheets ({len(hidden)}): {hidden}")

def export_graph_pdf(token, drive_id, item_id):
    r = graph_request(
        'GET',
        f"/drives/{drive_id}/items/{item_id}/content?format=pdf",
        token,
        headers={'Authorization': f'Bearer {token}', 'Accept': 'application/pdf'},
        allow_redirects=True,
        timeout=300,
    )
    if not is_valid_pdf(r.content):
        raise RuntimeError(f"Graph PDF export ne valid PDF nahi diya: {r.content[:80]!r}")
    return r.content

def make_pdf_with_excel_online(pdf_path):
    """
    VBA jaisi rendering ke liye Microsoft Excel Online se PDF export.
    Original workbook ko touch nahi karte: temp copy banti hai, us par sheets hide/order set hote hain.
    """
    token = get_graph_access_token()
    if not token:
        raise RuntimeError(
            "Graph credentials configured nahi hain. GRAPH_ACCESS_TOKEN ya "
            "MS_GRAPH_CLIENT_ID + MS_GRAPH_REFRESH_TOKEN set karein."
        )

    drive_id, source_item_id, parent_id = resolve_graph_drive_item(token)
    temp_item_id = None
    try:
        print("📄 Excel Online ke liye temp workbook copy ban rahi hai...")
        temp_item_id = copy_graph_workbook(token, drive_id, source_item_id, parent_id)
        prepare_graph_target_sheets(token, drive_id, temp_item_id)
        time.sleep(2)
        print("🖨️ Excel Online PDF export ho raha hai...")
        pdf_bytes = export_graph_pdf(token, drive_id, temp_item_id)
        Path(pdf_path).write_bytes(pdf_bytes)
        print(f"✅ Excel Online PDF complete ({len(pdf_bytes)//1024} KB)")
    finally:
        if temp_item_id:
            delete_graph_item(token, drive_id, temp_item_id)


# ═══════════════════════════════════════════════════════
#  STEP 3B: LibreOffice fallback — IsVisible toggle
#  (Better approach: print area + page break aware)
# ═══════════════════════════════════════════════════════

SOFFICE_PORT = 2002

def start_soffice(profile_dir):
    profile_uri = Path(profile_dir).resolve().as_uri()
    env = os.environ.copy()
    env['SAL_USE_VCLPLUGIN'] = 'svp'
    env['HOME'] = profile_dir
    log_file = open(os.path.join(profile_dir, 'lo.log'), 'w')
    proc = subprocess.Popen([
        'soffice', '--headless', '--invisible', '--nocrashreport',
        '--nodefault', '--norestore', '--nologo', '--nofirststartwizard',
        f'-env:UserInstallation={profile_uri}',
        f'--accept=socket,host=localhost,port={SOFFICE_PORT};urp;'
    ], env=env, stdout=log_file, stderr=log_file, stdin=subprocess.DEVNULL)
    proc._log_file = log_file
    return proc

def connect_uno(retries=90, proc=None):
    import uno
    ctx_local = uno.getComponentContext()
    resolver = ctx_local.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver", ctx_local)
    for i in range(retries):
        if proc and proc.poll() is not None:
            raise Exception(f"soffice crashed (exit {proc.returncode})")
        try:
            ctx = resolver.resolve(
                f"uno:socket,host=localhost,port={SOFFICE_PORT};urp;StarOffice.ComponentContext")
            print(f"✅ LibreOffice connected ({i+1} tries)")
            return ctx
        except Exception as e:
            if i % 15 == 0:
                print(f"⏳ Waiting for LibreOffice... ({i+1}/{retries})")
            time.sleep(1)
    raise Exception("LibreOffice se connect nahi hua")

def prop(name, value):
    from com.sun.star.beans import PropertyValue
    p = PropertyValue(); p.Name = name; p.Value = value
    return p

def make_pdf_with_libreoffice(xlsx_path, pdf_path, all_sheet_names):
    """
    Non-target sheets ko hide karke sirf target sheets ki PDF banao.
    Sheets delete nahi karte — cross-references safe rehti hain.
    """
    profile_dir = tempfile.mkdtemp(prefix='lo_profile_')
    proc = start_soffice(profile_dir)
    try:
        ctx = connect_uno(retries=90, proc=proc)
        smgr = ctx.ServiceManager
        desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)

        file_url = Path(xlsx_path).resolve().as_uri()
        doc = desktop.loadComponentFromURL(file_url, "_blank", 0, (
            prop("Hidden", True),
            prop("MacroExecutionMode", 0),
            prop("UpdateDocMode", 1),  # formulas recalculate
        ))

        sheets   = doc.Sheets
        all_lo   = list(sheets.ElementNames)
        target_order = []
        for target in TARGET_SHEETS:
            for name in all_lo:
                if norm_sheet(name) == norm_sheet(target):
                    target_order.append(name)
                    break

        if not target_order:
            raise Exception("LibreOffice export ke liye koi target sheet match nahi hui")

        print(f"📊 LibreOffice ne ye sheets dekhi: {all_lo}")

        try:
            doc.CurrentController.setActiveSheet(sheets.getByName(target_order[0]))
            print(f"✅ Active sheet target par set: {target_order[0]}")
        except Exception as ex:
            print(f"⚠️ Active sheet set nahi hui: {ex}")

        hidden_ok = []
        hide_fail = []
        for name in all_lo:
            is_target = norm_sheet(name) in TARGET_LOOKUP
            s = sheets.getByName(name)
            if is_target:
                # Target sheet — visible rakho
                try:
                    s.IsVisible = True
                except Exception:
                    pass
            else:
                # Non-target — hide karo
                try:
                    s.IsVisible = False
                    # Verify
                    if not s.IsVisible:
                        hidden_ok.append(name)
                    else:
                        hide_fail.append(name)
                except Exception as ex:
                    hide_fail.append(f"{name}({ex})")

        print(f"🙈 Hidden OK ({len(hidden_ok)}): {hidden_ok}")
        if hide_fail:
            print(f"⚠️ Hide FAILED ({len(hide_fail)}): {hide_fail}")

        # Verify visible sheets
        visible_now = [n for n in all_lo if sheets.getByName(n).IsVisible]
        print(f"👁️ Visible sheets jo PDF mein aayengi ({len(visible_now)}): {visible_now}")

        out_url = Path(pdf_path).resolve().as_uri()
        print("🖨️ PDF export ho raha hai...")
        t0 = time.time()
        doc.storeToURL(out_url, (prop("FilterName", "calc_pdf_Export"),))
        print(f"✅ PDF complete ({time.time()-t0:.1f}s)")
        doc.close(False)

    finally:
        try: proc.terminate(); proc.wait(timeout=8)
        except Exception:
            try: proc.kill(); proc.wait(timeout=5)
            except Exception: pass
        try:
            if hasattr(proc, '_log_file'): proc._log_file.close()
        except Exception: pass
        try:
            subprocess.run(['pkill','-9','-f','soffice'], timeout=5,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception: pass
        shutil.rmtree(profile_dir, ignore_errors=True)
        print("✅ Cleanup complete")


def make_block_pdf(xlsx_path, pdf_path, all_sheet_names):
    engine = PDF_ENGINE or 'auto'
    graph_engines = {'auto', 'graph', 'excel', 'excel-online', 'excel_online'}

    if engine in graph_engines:
        try:
            make_pdf_with_excel_online(pdf_path)
            return 'excel-online'
        except Exception as ex:
            if engine != 'auto':
                raise
            print(f"⚠️ Excel Online PDF export skip/fail: {ex}")
            print("↪️ LibreOffice fallback use ho raha hai.")
    elif engine not in {'libreoffice', 'lo'}:
        raise RuntimeError("PDF_ENGINE valid nahi hai. Use: auto, graph, excel-online, libreoffice")

    make_pdf_with_libreoffice(xlsx_path, pdf_path, all_sheet_names)
    return 'libreoffice'


# ═══════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════

def main():
    if not ONEDRIVE_URL:
        print("❌ ONEDRIVE_URL not set"); sys.exit(1)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        meta, xlsx_data = get_file_metadata_and_download(ONEDRIVE_URL)
        xlsx_path = tmp / 'workbook.xlsx'
        xlsx_path.write_bytes(xlsx_data)

        pdf_name, all_sheet_names = get_pdf_name_and_sheets(xlsx_data)

        pdf_out = tmp / 'output.pdf'
        engine_used = make_block_pdf(xlsx_path, pdf_out, all_sheet_names)

        final_pdf = DATA_DIR / f"{pdf_name}.pdf"
        shutil.copy(str(pdf_out), str(final_pdf))
        shutil.copy(str(pdf_out), str(DATA_DIR / 'block-report.pdf'))
        (DATA_DIR / 'block-pdf-name.txt').write_text(pdf_name)
        print(f"🎉 Done! PDF saved: {final_pdf.name} ({engine_used})")

if __name__ == '__main__':
    main()
    sys.stdout.flush()
    sys.stderr.flush()
    os._exit(0)
